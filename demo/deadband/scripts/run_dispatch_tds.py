#!/usr/bin/env python3
"""
Run a deadband-demo dispatch interval through ANDES TDS.

The primary entrypoint is a dispatch JSON produced from the deadband demo
workflow, but the script can still recompute a dispatch from AMS when
``--dispatch-json`` is omitted.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
CASES = ROOT / "cases"
RESULTS = ROOT / "results"


def find_workspace(root: Path) -> Path | None:
    """
    Locate a workspace containing sibling ``andes`` and ``ams`` source trees.

    This keeps the demo runnable both inside the historical
    ``openandes/demo/demo/deadband`` layout and from a standalone export such as
    the ``deadband2`` repository.
    """
    candidates: list[Path] = []

    env_workspace = os.environ.get("OPENANDES_WORKSPACE")
    if env_workspace:
        candidates.append(Path(env_workspace).expanduser())

    resolved = root.resolve()
    for parent in (resolved, *resolved.parents):
        candidates.append(parent)
        candidates.append(parent / "openandes")

    seen: set[Path] = set()
    for candidate in candidates:
        try:
            candidate = candidate.resolve()
        except FileNotFoundError:
            continue

        if candidate in seen:
            continue
        seen.add(candidate)

        if (candidate / "andes").exists() and (candidate / "ams").exists():
            return candidate

    return None


WORKSPACE = find_workspace(ROOT)

# Prefer the local ANDES/AMS source trees so demo scripts pick up the
# compatibility patches in this workspace without requiring PYTHONPATH.
if WORKSPACE is not None:
    for src_root in (WORKSPACE / "andes", WORKSPACE / "ams"):
        src_str = str(src_root)
        if src_root.exists() and src_str not in sys.path:
            sys.path.insert(0, src_str)

import andes
from andes.thirdparty.npfunc import safe_div

DEFAULT_OPF_CASE = CASES / "IL200_opf2.xlsx"
DEFAULT_DYN_CASE = CASES / "IL200_dyn_db2.xlsx"
DEFAULT_CURVE_FILE = CASES / "CurveInterp.csv"
DEFAULT_STABLE_DYN_CASE = CASES / "IL200_dyn_db2_stable.xlsx"
DEFAULT_WIND_PREFIXES = ("WT_",)
DEFAULT_SOLAR_PREFIXES = ("PV_",)


@dataclass
class DispatchRecord:
    hour: int
    dispatch: int
    load: float
    wind: float
    solar: float
    gen: list
    pg: list
    qg: list
    pd: list
    bus: list
    vBus: list
    aBus: list
    converged: bool = True
    obj: float = float("nan")

    @property
    def label(self) -> str:
        return f"h{self.hour}d{self.dispatch}"

    @classmethod
    def from_json(cls, path: Path) -> "DispatchRecord":
        return cls(**json.loads(path.read_text()))


def normalize_prefixes(prefixes: Iterable[str] | None, defaults: tuple[str, ...]) -> tuple[str, ...]:
    if prefixes is None:
        return defaults

    items = tuple(prefix for prefix in prefixes if prefix)
    return items or defaults


def adapt_dyn_case(src: Path, dst: Path) -> Path:
    """
    Create a stable-style copy of the legacy dynamic case.

    The legacy deadband case uses ``PVD2`` / ``ESD2``. For migration, rename
    them to ``PVD1`` / ``ESD1`` and add the optional ``fdbdu`` column
    explicitly. A default of ``0.017`` preserves the historical upper
    deadband used by the legacy bi-directional models when the column was
    omitted.
    """

    def rename_sheet(wb: openpyxl.Workbook, old: str, new: str) -> None:
        if new in wb.sheetnames and old in wb.sheetnames:
            del wb[new]
        if old in wb.sheetnames:
            wb[old].title = new

    def ensure_column(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        after: str,
        name: str,
        value: float,
    ) -> None:
        headers = [cell.value for cell in ws[1]]
        if name in headers or after not in headers:
            return

        insert_idx = headers.index(after) + 2
        ws.insert_cols(insert_idx)
        ws.cell(row=1, column=insert_idx, value=name)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=insert_idx, value=value)

    wb = openpyxl.load_workbook(src)
    rename_sheet(wb, "PVD2", "PVD1")
    rename_sheet(wb, "ESD2", "ESD1")

    for sheet in ("PVD1", "ESD1"):
        if sheet in wb.sheetnames:
            ensure_column(wb[sheet], after="fdbd", name="fdbdu", value=0.017)

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    return dst


def make_sp(opf_case: Path) -> Any:
    import ams

    return ams.load(str(opf_case), setup=True, no_output=True, default_config=True)


def load_curve(curve_file: Path) -> pd.DataFrame:
    return pd.read_csv(curve_file)


def compute_dispatch(
    hour: int,
    dispatch: int,
    curve: pd.DataFrame,
    opf_case: Path,
    dispatch_interval: int,
) -> DispatchRecord:
    """
    Recompute one ACOPF dispatch interval using the demo notebook logic.
    """
    r0 = hour * 3600 + dispatch * dispatch_interval
    r1 = r0 + dispatch_interval

    sp = make_sp(opf_case)

    pq_idx = sp.PQ.idx.v
    p0 = sp.PQ.p0.v.copy()
    q0 = sp.PQ.q0.v.copy()
    stg = sp.StaticGen.get_all_idxes()
    stg_w2t, stg_pv, _ = sp.StaticGen.find_idx(
        keys="gentype",
        values=["W2", "PV", "ES"],
        allow_all=True,
    )
    p0_w2t = sp.StaticGen.get(src="p0", attr="v", idx=stg_w2t)
    p0_pv = sp.StaticGen.get(src="p0", attr="v", idx=stg_pv)

    load = curve["Load"].iloc[r0:r1].values.mean()
    sp.PQ.set(src="p0", idx=pq_idx, attr="v", value=load * p0)
    sp.PQ.set(src="q0", idx=pq_idx, attr="v", value=load * q0)

    psum = sp.PQ.p0.v.sum()
    solar = curve["PV"].iloc[r0:r1].values.mean()
    wind = curve["Wind"].iloc[r0:r1].values.mean()

    wind_sum = wind * p0_w2t.sum()
    solar_sum = solar * p0_pv.sum()
    if wind_sum + solar_sum > psum:
        dgen = wind_sum + solar_sum - psum
        dwind = dgen / (wind_sum + solar_sum) * wind_sum
        dsolar = dgen / (wind_sum + solar_sum) * solar_sum
        wind = safe_div(wind_sum - 1.05 * dwind, wind_sum)
        solar = safe_div(solar_sum - 1.05 * dsolar, solar_sum)

    sp.StaticGen.set(src="p0", idx=stg_w2t, attr="v", value=wind * p0_w2t)
    sp.StaticGen.set(src="p0", idx=stg_pv, attr="v", value=solar * p0_pv)

    pmax = sp.StaticGen.get(src="pmax", attr="v", idx=stg).copy()
    sp.StaticGen.set(src="pmax", idx=stg, attr="v", value=0.995 * pmax)

    sp.ACOPF.config.update(verbose=0, out_all=0)
    sp.ACOPF.update()
    sp.ACOPF.run()

    sp.StaticGen.set(src="pmax", idx=stg, attr="v", value=pmax)

    return DispatchRecord(
        hour=hour,
        dispatch=dispatch,
        load=float(load),
        wind=float(wind),
        solar=float(solar),
        gen=sp.ACOPF.pg.get_all_idxes(),
        pg=sp.ACOPF.pg.v.tolist(),
        qg=sp.ACOPF.qg.v.tolist(),
        pd=sp.ACOPF.pd.v.tolist(),
        bus=sp.ACOPF.vBus.get_all_idxes(),
        vBus=sp.ACOPF.vBus.v.tolist(),
        aBus=sp.ACOPF.aBus.v.tolist(),
        converged=bool(sp.ACOPF.converged),
        obj=float(sp.ACOPF.obj.v),
    )


def build_andes_link(sa: andes.system.System) -> pd.DataFrame:
    """
    Build the minimal generator link table needed for AGC without AMS.
    """
    stg_idx = sa.StaticGen.get_all_idxes()
    dg_idx = sa.DG.find_idx(keys="gen", values=stg_idx, allow_none=True)
    rg_idx = sa.RenGen.find_idx(keys="gen", values=stg_idx, allow_none=True)

    syg_idx = sa.SynGen.get_all_idxes()
    syg_gen = sa.SynGen.get(src="gen", attr="v", idx=syg_idx)
    gov_idx = sa.TurbineGov.find_idx(keys="syn", values=syg_idx, allow_none=True)
    gov_map = {int(gen): gov for gen, gov in zip(syg_gen, gov_idx)}

    gammap = np.ones(len(stg_idx), dtype=float)
    for i, dg in enumerate(dg_idx):
        if dg:
            gammap[i] = float(sa.DG.get(src="gammap", attr="v", idx=dg))
        elif rg_idx[i]:
            gammap[i] = float(sa.RenGen.get(src="gammap", attr="v", idx=rg_idx[i]))

    link = pd.DataFrame(
        {
            "stg_idx": stg_idx,
            "gov_idx": [gov_map.get(int(idx)) for idx in stg_idx],
            "dg_idx": dg_idx,
            "rg_idx": rg_idx,
            "gammap": gammap,
        }
    )
    link["has_gov"] = link["gov_idx"].notna().astype(int)
    link["has_dg"] = link["dg_idx"].notna().astype(int)
    link["has_rg"] = link["rg_idx"].notna().astype(int)
    link[["agov", "adg", "arg"]] = 0.0
    return link


def pvd1_gen_subsets(
    sa: andes.system.System,
    wind_prefixes: Iterable[str],
    solar_prefixes: Iterable[str],
) -> tuple[list[int], list[int]]:
    """
    Split PVD1 devices into wind and PV subsets using idx/name prefixes.
    """
    wind_prefixes = tuple(wind_prefixes)
    solar_prefixes = tuple(solar_prefixes)

    names = getattr(sa.PVD1, "name", None)
    name_values = names.v if names is not None else [None] * sa.PVD1.n

    wind: list[int] = []
    solar: list[int] = []

    for idx, name, gen in zip(sa.PVD1.idx.v, name_values, sa.PVD1.gen.v):
        labels = [str(idx)]
        if name is not None:
            labels.append(str(name))

        if any(label.startswith(prefix) for label in labels for prefix in wind_prefixes):
            wind.append(int(gen))
        if any(label.startswith(prefix) for label in labels for prefix in solar_prefixes):
            solar.append(int(gen))

    if not wind or not solar:
        sample = ", ".join(map(str, sa.PVD1.idx.v[:10]))
        raise ValueError(
            "Unable to classify PVD1 devices from prefixes. "
            f"wind_prefixes={wind_prefixes}, solar_prefixes={solar_prefixes}, "
            f"sample_idx=[{sample}]"
        )

    return wind, solar


def validate_curve_window(curve: pd.DataFrame, dispatch_record: DispatchRecord, duration_seconds: int) -> None:
    r0 = dispatch_record.hour * 3600 + dispatch_record.dispatch * duration_seconds
    r1 = r0 + duration_seconds
    if r1 > len(curve):
        raise ValueError(
            f"Curve data only has {len(curve)} samples but {dispatch_record.label} "
            f"needs samples [{r0}, {r1})."
        )


def resolve_initial_profile(
    curve: pd.DataFrame,
    dispatch_record: DispatchRecord,
    duration_seconds: int,
    init_mode: str,
) -> tuple[float, float, float]:
    """
    Resolve the TDS starting load / wind / solar point.

    ``dispatch`` preserves the historical behavior of initializing from the
    dispatch-interval average. ``first`` uses the first curve sample in the
    interval.
    """
    if init_mode == "dispatch":
        return (
            float(dispatch_record.load),
            float(dispatch_record.wind),
            float(dispatch_record.solar),
        )

    r0 = dispatch_record.hour * 3600 + dispatch_record.dispatch * duration_seconds
    return (
        float(curve["Load"].iloc[r0]),
        float(curve["Wind"].iloc[r0]),
        float(curve["PV"].iloc[r0]),
    )


def run_tds(
    dispatch_record: DispatchRecord,
    curve: pd.DataFrame,
    dyn_case: Path,
    duration_seconds: int,
    agc_interval: int,
    kp: float,
    ki: float,
    wind_prefixes: Iterable[str],
    solar_prefixes: Iterable[str],
    init_mode: str = "first",
) -> tuple[np.ndarray, np.ndarray]:
    """
    Run one dispatch interval and return time and ACE frequency deviation.
    """
    validate_curve_window(curve, dispatch_record, duration_seconds)

    sa = andes.load(str(dyn_case), setup=False, no_output=True, default_config=True)

    # Record a direct frequency trace for plotting.
    sa.add("Output", dict(model="ACEc", varname="f"))

    sa.setup()

    link = build_andes_link(sa)

    pq_idx = sa.PQ.idx.v
    stg = sa.StaticGen.get_all_idxes()
    stg_w2t, stg_pv = pvd1_gen_subsets(sa, wind_prefixes, solar_prefixes)
    p0_w2t = sa.StaticGen.get(src="p0", attr="v", idx=stg_w2t)
    p0_pv = sa.StaticGen.get(src="p0", attr="v", idx=stg_pv)
    pvd1_w2t = sa.PVD1.find_idx(keys="gen", values=stg_w2t)
    pvd1_pv = sa.PVD1.find_idx(keys="gen", values=stg_pv)

    sap0 = sa.PQ.p0.v.copy()
    saq0 = sa.PQ.q0.v.copy()

    sa.StaticGen.set(src="p0", idx=dispatch_record.gen, attr="v", value=dispatch_record.pg)
    sa.Bus.set(src="v0", idx=dispatch_record.bus, attr="v", value=dispatch_record.vBus)
    sa.Bus.set(src="a0", idx=dispatch_record.bus, attr="v", value=dispatch_record.aBus)

    pv_bus = sa.PV.bus.v
    slack_bus = sa.Slack.bus.v
    v_pv = sa.Bus.get(src="v0", attr="v", idx=pv_bus)
    a_slack = sa.Bus.get(src="a0", attr="v", idx=slack_bus)
    sa.PV.set(src="v0", idx=sa.PV.idx.v, attr="v", value=v_pv)
    sa.Slack.set(src="a0", idx=sa.Slack.idx.v, attr="v", value=a_slack)

    stg_on_uid = np.where(np.array(dispatch_record.pg) > 1e-4)[0]
    stg_on = np.array([1 if uid in stg_on_uid else 0 for uid in range(len(stg))])
    sn = sa.StaticGen.get(src="Sn", attr="v", idx=stg)
    bf = stg_on * sn / (stg_on * sn).sum()

    sa.PQ.config.p2p = 1
    sa.PQ.config.q2q = 1
    sa.PQ.config.p2z = 0
    sa.PQ.config.q2z = 0
    sa.PQ.pq2z = 0

    sa.TDS.config.criteria = 0
    sa.TDS.config.no_tqdm = True

    init_load, init_wind, init_solar = resolve_initial_profile(
        curve=curve,
        dispatch_record=dispatch_record,
        duration_seconds=duration_seconds,
        init_mode=init_mode,
    )

    sa.PQ.set(src="p0", idx=pq_idx, attr="v", value=init_load * sap0)
    sa.PQ.set(src="q0", idx=pq_idx, attr="v", value=init_load * saq0)
    sa.StaticGen.set(src="p0", idx=stg_w2t, attr="v", value=init_wind * p0_w2t)
    sa.StaticGen.set(src="p0", idx=stg_pv, attr="v", value=init_solar * p0_pv)

    sa.PFlow.run()
    if sa.exit_code != 0:
        raise RuntimeError(f"PFlow failed with exit_code={sa.exit_code}")

    _ = sa.TDS.init()
    if sa.exit_code != 0:
        raise RuntimeError(f"TDS init failed with exit_code={sa.exit_code}")

    pext_max = 999 * np.ones(sa.DG.n)
    if hasattr(sa, "ESD1") and sa.ESD1.n:
        ess_uid = sa.DG.idx2uid(sa.ESD1.idx.v)
        pext_max[ess_uid] = 999

    ace_integral = 0.0
    ace_raw = 0.0
    r0 = dispatch_record.hour * 3600 + dispatch_record.dispatch * duration_seconds
    t_snapshots = [0.0]
    f_snapshots = [float((sa.ACEc.f.v[0] - 1.0) * sa.config.freq)]

    for t in range(duration_seconds):
        for col, has_col in (("agov", "has_gov"), ("adg", "has_dg"), ("arg", "has_rg")):
            link[col] = ace_raw * bf * link[has_col] * link["gammap"]

        if t % agc_interval == 0 and t > 0:
            agov_to_set = {
                gov: agov for gov, agov in zip(link["gov_idx"], link["agov"]) if pd.notna(gov)
            }
            if agov_to_set:
                gov_idx = list(agov_to_set.keys())
                paux0_raw = np.array(list(agov_to_set.values()))
                gov_syn = sa.TurbineGov.get(src="syn", attr="v", idx=gov_idx)
                gov_gen = sa.SynGen.get(src="gen", attr="v", idx=gov_syn)
                gov_pmax = sa.StaticGen.get(src="pmax", attr="v", idx=gov_gen)
                gov_pmin = sa.StaticGen.get(src="pmin", attr="v", idx=gov_gen)
                gov_pref0 = sa.TurbineGov.get(src="pref0", attr="v", idx=gov_idx)
                gov_up = np.maximum(0.0, gov_pmax - gov_pref0)
                gov_dn = np.minimum(0.0, gov_pmin - gov_pref0)
                # Preserve the AGC command sign while respecting generator headroom.
                paux0 = np.where(
                    paux0_raw >= 0.0,
                    np.minimum(paux0_raw, gov_up),
                    np.maximum(paux0_raw, gov_dn),
                )
                sa.TurbineGov.set(src="paux0", idx=gov_idx, attr="v", value=paux0)

            adg_to_set = {dg: adg for dg, adg in zip(link["dg_idx"], link["adg"]) if pd.notna(dg)}
            if adg_to_set:
                dg_idx = list(adg_to_set.keys())
                pext0_raw = np.array(list(adg_to_set.values()))
                dg_uids = sa.DG.idx2uid(dg_idx)
                pext0 = np.minimum(pext0_raw, pext_max[dg_uids])
                sa.DG.set(src="Pext0", idx=dg_idx, attr="v", value=pext0)

        if t > 0:
            kload = curve["Load"].iloc[r0 + t]
            sa.PQ.set(src="Ppf", idx=sa.PQ.idx.v, attr="v", value=kload * sap0)
            sa.PQ.set(src="Qpf", idx=sa.PQ.idx.v, attr="v", value=kload * saq0)

            wind = curve["Wind"].iloc[r0 + t]
            sa.PVD1.set(src="pref0", idx=pvd1_w2t, attr="v", value=wind * p0_w2t)

            solar = curve["PV"].iloc[r0 + t]
            sa.PVD1.set(src="pref0", idx=pvd1_pv, attr="v", value=solar * p0_pv)

            sa.TDS.config.tf = t
            sa.TDS.run()
            t_snapshots.append(float(sa.dae.t))
            f_snapshots.append(float((sa.ACEc.f.v[0] - 1.0) * sa.config.freq))

            ace_sum = sa.ACEc.ace.v.sum()
            ace_raw = -(kp * ace_sum + ki * ace_integral)
            ace_integral = ace_integral + ace_sum

        if sa.exit_code != 0:
            raise RuntimeError(f"TDS failed at t={t}s with exit_code={sa.exit_code}")

    t = np.asarray(sa.dae.ts.t).reshape(-1)
    f_pu = np.asarray(sa.dae.ts.get_data(sa.ACEc.f, a=[0])).reshape(-1)
    n = min(len(t), len(f_pu))
    if n > 0:
        f_dev_hz = (f_pu[:n] - 1.0) * sa.config.freq
        return t[:n], f_dev_hz

    return np.asarray(t_snapshots), np.asarray(f_snapshots)


def save_outputs(
    t: np.ndarray,
    f_dev_hz: np.ndarray,
    dispatch_record: DispatchRecord,
    out_dir: Path,
    label: str | None = None,
) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{label or dispatch_record.label}_frequency"

    csv_path = out_dir / f"{stem}.csv"
    png_path = out_dir / f"{stem}.png"

    pd.DataFrame({"time_s": t, "freq_dev_hz": f_dev_hz}).to_csv(csv_path, index=False)

    fig, ax = plt.subplots(figsize=(9, 4.8))
    ax.plot(t, f_dev_hz, color="#0f5c78", linewidth=1.4)
    ax.axhline(0.0, color="#777777", linewidth=0.8, linestyle="--")
    ax.set_title(f"Deadband Demo Frequency Deviation ({label or dispatch_record.label})")
    ax.set_xlabel("Time [s]")
    ax.set_ylabel("Frequency deviation [Hz]")
    ax.grid(True, alpha=0.25)
    fig.tight_layout()
    fig.savefig(png_path, dpi=180)
    plt.close(fig)

    return csv_path, png_path


def write_dispatch_json(
    dispatch_record: DispatchRecord,
    out_dir: Path,
    label: str | None = None,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{label or dispatch_record.label}_dispatch.json"
    path.write_text(json.dumps(dispatch_record.__dict__, indent=2))
    return path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dispatch-json", type=Path, default=None,
                        help="Existing dispatch JSON to replay through TDS.")
    parser.add_argument("--hour", type=int, default=13,
                        help="Dispatch hour used when recomputing from AMS.")
    parser.add_argument("--dispatch", type=int, default=2,
                        help="Dispatch interval used when recomputing from AMS.")
    parser.add_argument("--label", type=str, default=None,
                        help="Output label. Defaults to h<hour>d<dispatch>.")
    parser.add_argument("--opf-case", type=Path, default=DEFAULT_OPF_CASE)
    parser.add_argument("--dyn-case", type=Path, default=DEFAULT_DYN_CASE)
    parser.add_argument("--stable-dyn-case", type=Path, default=DEFAULT_STABLE_DYN_CASE)
    parser.add_argument("--curve-file", type=Path, default=DEFAULT_CURVE_FILE)
    parser.add_argument("--results-dir", type=Path, default=RESULTS)
    parser.add_argument("--duration-seconds", type=int, default=900)
    parser.add_argument("--agc-interval", type=int, default=4)
    parser.add_argument("--kp", type=float, default=0.05)
    parser.add_argument("--ki", type=float, default=0.0625)
    parser.add_argument("--init-mode", choices=("dispatch", "first"),
                        default="first",
                        help="TDS initialization profile: dispatch average or a curve sample.")
    parser.add_argument("--wind-prefix", action="append", default=None,
                        help="PVD1 idx/name prefix for wind units. Repeatable.")
    parser.add_argument("--solar-prefix", action="append", default=None,
                        help="PVD1 idx/name prefix for solar units. Repeatable.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    andes.config_logger(stream_level=30)

    curve = load_curve(args.curve_file)
    dyn_case = adapt_dyn_case(args.dyn_case, args.stable_dyn_case)
    wind_prefixes = normalize_prefixes(args.wind_prefix, DEFAULT_WIND_PREFIXES)
    solar_prefixes = normalize_prefixes(args.solar_prefix, DEFAULT_SOLAR_PREFIXES)

    if args.dispatch_json is not None:
        dispatch_record = DispatchRecord.from_json(args.dispatch_json)
    else:
        import ams

        ams.config_logger(stream_level=50)
        dispatch_record = compute_dispatch(
            args.hour,
            args.dispatch,
            curve,
            args.opf_case,
            args.duration_seconds,
        )

    if not dispatch_record.converged:
        raise RuntimeError(f"Dispatch {dispatch_record.label} did not converge")

    label = args.label or dispatch_record.label
    t, f_dev_hz = run_tds(
        dispatch_record=dispatch_record,
        curve=curve,
        dyn_case=dyn_case,
        duration_seconds=args.duration_seconds,
        agc_interval=args.agc_interval,
        kp=args.kp,
        ki=args.ki,
        wind_prefixes=wind_prefixes,
        solar_prefixes=solar_prefixes,
        init_mode=args.init_mode,
    )
    dispatch_json = write_dispatch_json(dispatch_record, args.results_dir, label=label)
    csv_path, png_path = save_outputs(t, f_dev_hz, dispatch_record, args.results_dir, label=label)

    print(f"dispatch_json={dispatch_json}")
    print(f"freq_csv={csv_path}")
    print(f"freq_plot={png_path}")
    print(f"freq_dev_min_hz={float(f_dev_hz.min())}")
    print(f"freq_dev_max_hz={float(f_dev_hz.max())}")
    print(f"samples={len(t)}")


if __name__ == "__main__":
    main()
